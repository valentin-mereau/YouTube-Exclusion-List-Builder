import time
import openpyxl
import tkinter as tk

from tkinter import ttk, filedialog, PhotoImage, messagebox
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from lingua import LanguageDetectorBuilder


class MainApp(tk.Tk):
    """The MainApp class is the core of the application with all the layout set and the logic under its feature.

    The class inherits from tinker.TK which construct a toplevel Tk widget.

    Attributes
    ----------
    title : tkinter.Wm
        the title of the application displayed at the top of the window
    resizable : tkinter.Wm
        control the possibility to resize the app window on the x and y axes
    excel_file_path : tkinter.StringVar()
        contain the path of the uploaded file
    icon_image : tkinter.PhotoImage
        excel icon stored in base 64 displayed next to the file name
    api_key : str
        define the validity of the user's token
    channel_number : int
        count the number of channel to process on the uploaded file
    processed_channel : int
        count the channels already processed
    workbook : openpyxl.reader.excel
        loaded Excel file
    stop_and_save_state : bool
        indicate when the button 'Stop & Save' is used

    Methods
    -------
    browse_file()
        ask user to upload file and update interface
    help_window()
        display help instructions
    process_channels()
        start the process to check channels and update interface
    stop_and_save()
        stop process from user action
    is_valid_youtube_token()
        check if user's token is valid
    verify_excel_template()
        verifies if uploaded file is conform to template
    youtube_checker()
        processes channels
    end_process()
        reset the app
    """
    def __init__(self):
        """Inits MainApp with all the attributes necessary to the logic of the application and the gui layout"""
        super().__init__()

        self.title("YouTube Made For Kid Checker")
        self.resizable(False, False)
        self.excel_file_path = tk.StringVar()
        self.icon_image = PhotoImage(data=('iVBORw0KGgoAAAANSUhEUgAAAB4AAAAeCAYAAAA7MK6iAAAACXBIWXMAAAsTAAALEwEAmpw'
                                           'YAAAB+ElEQVR4nGNgGAWDCQj3h6qLTQpywYc191U6ax6ud8GJD9Wak2SpyORwNbEJwX/EJo'
                                           'b8x4c1DtY+1jxU/x8vPlyfRLTF4v3BDoQsJdrig/UNoxaDACNvvpUld4GFCzLmr3EqFu7w/'
                                           'Y+ORfsCqRPUPMXmaTxFFv+Jx5Y/+Zo9jsGwzNqMfTIbMo/hxZtzijEs5i40byDNYov/PA2u'
                                           'cCyxOuWJ5JrU/3jx6tQGgharN/r9V6zzhPO1mgP+S1e70N7i9p1z/7/49Oa/WIXDf8eJKf9'
                                           '//P7136gjnPYWi5bb/3/49tn/6s2T/x+/d/F/28459AlqniKL/3GLqsE+vfv68X+RMjv6WZ'
                                           'y+vPn/x++f/9978wQcAujy4u5acCw4Jeix0LSg/wRwPUGL5Ws9/r/+/P5/xLyy/6ceXP7fv'
                                           'WcBAYsDnxCyWHBqIGEfrzq36/+Zh1f/8xZb/veYmvn/998//y17YmlrMW+x5X+d1mCU7KPV'
                                           'EvhfptqV9j7mIQJTxWIeEotM3jyzn+JuWsdgWLTDe49Ir98xfFii3aeQ+Eoi3rhYNFjvPzo'
                                           'W99H+L+GhBcdqqXaP1dPs/+PFqfYk1Mfu2g7IFuDCoxarUxrUUp4aahLuWn+oEtRp9olEWw'
                                           'yzXMxd0wUfVkuycVZNtXfBidNtzUiydBQwUAkAAEux+AOOo/0gAAAAAElFTkSuQmCC'))
        self.api_key = "empty"
        self.channel_number = 0
        self.processed_channel = 0
        self.workbook = None
        self.stop_and_save_state = False

        self.frame_main = Container(self, column_number=0, row_number=0)

        # notebook settings
        self.notebook = ttk.Notebook(self.frame_main)
        self.notebook.grid(row=0, column=0, sticky='nswe')

        # tab1
        self.tab1_container = Container(self.notebook, column_number=(0, 1, 2), row_number=(0, 1, 2, 3), uniform_type='a')
        self.notebook.add(self.tab1_container, text="Excel File")

        self.lbl_file_uploaded = ttk.Label(self.tab1_container, text="No file uploaded", anchor=tk.CENTER)
        self.lbl_file_uploaded.grid(row=0, column=0, sticky="swe", rowspan=2, columnspan=3)

        self.lbl_yb_channel_count = ttk.Label(self.tab1_container, anchor=tk.CENTER)
        self.lbl_yb_channel_count.grid(row=3, column=0, sticky="nswe")

        self.charging_bar = ChargingBar(self.tab1_container)

        self.btn_upload = ttk.Button(self.tab1_container, text=f"Upload excel file", command=self.browse_file)
        self.btn_upload.grid(row=3, column=2)

        self.btn_process = ttk.Button(self.tab1_container, text=f"Process channels", command=self.process_channels, state=tk.DISABLED)
        self.btn_process.grid(row=3, column=1, padx=5)

        # tab2
        self.tab2_container = Container(self.notebook, column_number=(0, 1), row_number=(0, 1, 2), uniform_type='a')
        self.notebook.add(self.tab2_container, text="API Token")

        self.lbl_entry_title = ttk.Label(self.tab2_container, text="Enter valid Youtube API token:")
        self.lbl_entry_title.grid(row=0, column=0, columnspan=2)

        self.api_entry = ttk.Entry(self.tab2_container, width=39)
        self.api_entry.grid(row=1, column=0, sticky="ns", columnspan=2, pady=5)

        self.btn_help = ttk.Button(self.tab2_container, text="How to get a token ?", command=self.help_window)
        self.btn_help.grid(row=2, column=1)

    def browse_file(self):
        """The browsing function will prompt the user to select and upload file.

        The file uploaded will be set to an attribute and verified by external function.
        If the file is valid, it will be displayed on the interface and the user will have the possibility to process it.

        Only .xlsx file are allow as we are working with openpyxl.
        """
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            # if file was provided set the file path and verify if the file match the template
            self.excel_file_path.set(filepath)
            if self.verify_excel_template():
                # if the file is valid, display file name and icon
                self.lbl_file_uploaded.config(text=filepath.split('/')[-1], image=self.icon_image, compound='left')
                self.lbl_yb_channel_count.config(text=f'{self.channel_number - self.processed_channel} channels')
                # activate the process button
                self.btn_process.config(state=tk.NORMAL)
            else:
                # It's only useful when a first valid file is uploaded and a second invalid file is uploaded
                # Reset the label text and image
                self.lbl_file_uploaded.config(text="No file uploaded", image="")
                # Reset channel counter
                self.lbl_yb_channel_count.config(text="")
                # Rest process button
                self.btn_process.config(state=tk.DISABLED)

    def help_window(self):
        """This function will call an instance of the HelpWindow class and show the instruction how to get a token."""
        return HelpWindow(master=self)

    def process_channels(self):
        """Logic structure used when 'process channel' button is triggerd.
        The token validity will be first verified.
        If the token is valid:
            The upload button and the api entry will be disabled to avoid any conflicting changes during the process.
            The charging bar and the save&quit feature will be shown.
            The function to process the channels will start.
        """
        self.is_valid_youtube_token()  # initiate verification of api toke

        # display error message box if token is invalid or empty
        if self.api_key == "invalid":
            messagebox.showinfo(title="Message Box", message="Token invalid", icon='error')
        elif self.api_key == "empty":
            messagebox.showinfo(title="Message Box", message="Token empty", icon='error')
        else:
            # when token is valid, disable token entry and upload button to avoid any changes during process
            self.api_entry.config(state=tk.DISABLED)
            self.btn_upload.config(state=tk.DISABLED)
            # display the charging bar use to show progress
            self.charging_bar.show_bar()
            # alter process channel button to have Stop & Save feature
            self.btn_process.config(text="Stop & Save", command=self.stop_and_save)
            # start processing channels
            self.youtube_checker()

    def stop_and_save(self):
        """Change the state of the attribute stop_and_save_state in order to exit the processing channel loop."""
        self.stop_and_save_state = True

    def is_valid_youtube_token(self):
        """Verify the validity of the token by sending a request to the Google api under a try statement."""
        if self.api_entry.get():
            # if there is text in api entry
            try:
                # try to send request to api with token provided
                youtube = build('youtube', 'v3', developerKey=self.api_entry.get())
                youtube.videos().list(part='id', id='VIDEO_ID').execute()
                # set api key to valid if the request is successful
                self.api_key = "valid"

            except HttpError:
                # if it catches an error the token is not valid
                self.api_key = "invalid"

    def verify_excel_template(self):
        """Verifies if the file uploaded is matching the template and count channels to process inside the file"""
        # load the file
        self.workbook = openpyxl.load_workbook(self.excel_file_path.get())

        # reset attributes for the case when user change the uploaded file
        self.channel_number = 0
        self.processed_channel = 0

        # if a result tab is present, verify header
        if ["Data", "Results"] == self.workbook.sheetnames and ["Placement", "Placement URL", "madeForKids", "Description"] == [cell.value for cell in self.workbook["Results"][1]]:
            # get the number of processed channels
            self.processed_channel = sum(1 for row in self.workbook["Results"].iter_rows(min_row=2, values_only=True) if any(row))
        # if data tab is present, verify header
        if "Data" in self.workbook.sheetnames and ["Placement", "Placement URL"] == [cell.value for cell in self.workbook["Data"][1]]:
            # get the number of channel in the data tab
            self.channel_number = sum(1 for row in self.workbook["Data"].iter_rows(min_row=2, values_only=True) if any(row))

            # return False and display error message box when template is matching but no data to process
            if not self.channel_number:
                messagebox.showinfo(title="Message Box", message="Empty file", icon='error')
                return False
            elif self.channel_number == self.processed_channel:
                messagebox.showinfo(title="Message Box", message="Channels already processed", icon='error')
                return False

            return True

        else:
            # return False and display error message box when template invalid
            messagebox.showinfo(title="Message Box", message="Template file incorrect", icon='error')
            return False

    def youtube_checker(self):
        """Main function to process the channels in the uploaded file"""
        def calculation_process_time(start_time, current_iter, max_iter):
            """Calculates an estimation of the time left to process all channels with basic math"""
            t_elapsed = time.time() - start_time
            t_estimated = (t_elapsed / current_iter) * max_iter
            time_left = t_estimated - t_elapsed

            if time_left >= 2 * 3600:  # 2 hours or more
                return f'{round(time_left / 3600)} hours left'
            elif time_left >= 3600:  # 1 hour or more
                return f'{round(time_left / 3600)} hour left'
            elif time_left >= 2 * 60:  # 2 minutes or more
                return f'{round(time_left / 60)} minutes left'
            elif time_left >= 60:  # 1 minute or more
                return f'{round(time_left / 60)} minute left'
            else:  # less than 1 minute
                return f'{round(time_left)} seconds left'

        def get_youtube_api_service(api_key):
            """Sets the header of api request with correct service, version and user token"""
            api_service_name = "youtube"
            api_version = "v3"
            return build(api_service_name, api_version, developerKey=api_key)

        def get_channel_properties(api_service, channel_url):
            """Gets channel properties"""
            topic_id = {"/m/04rlf": "Music (parent topic)", "/m/02mscn": "Christian music",
                        "/m/0ggq0m": "Classical music", "/m/01lyv": "Country", "/m/02lkt": "Electronic music",
                        "/m/0glt670": "Hip hop music", "/m/05rwpb": "Independent music", "/m/03_d0": "Jazz",
                        "/m/028sqc": "Music of Asia", "/m/0g293": "Music of Latin America", "/m/064t9": "Pop music",
                        "/m/06cqb": "Reggae", "/m/06j6l": "Rhythm and blues", "/m/06by7": "Rock music",
                        "/m/0gywn": "Soul music", "/m/0bzvm2": "Gaming (parent topic)", "/m/025zzc": "Action game",
                        "/m/02ntfj": "Action-adventure game", "/m/0b1vjn": "Casual game",
                        "/m/02hygl": "Music video game", "/m/04q1x3q": "Puzzle video game",
                        "/m/01sjng": "Racing video game", "/m/0403l3g": "Role-playing video game",
                        "/m/021bp2": "Simulation video game", "/m/022dc6": "Sports game",
                        "/m/03hf_rm": "Strategy video game", "/m/06ntj": "Sports (parent topic)",
                        "/m/0jm_": "American football", "/m/018jz": "Baseball", "/m/018w8": "Basketball",
                        "/m/01cgz": "Boxing", "/m/09xp_": "Cricket", "/m/02vx4": "Football", "/m/037hz": "Golf",
                        "/m/03tmr": "Ice hockey", "/m/01h7lh": "Mixed martial arts", "/m/0410tth": "Motorsport",
                        "/m/07bs0": "Tennis", "/m/07_53": "Volleyball", "/m/02jjt": "Entertainment (parent topic)",
                        "/m/09kqc": "Humor", "/m/02vxn": "Movies", "/m/05qjc": "Performing arts",
                        "/m/066wd": "Professional wrestling", "/m/0f2f9": "TV shows",
                        "/m/019_rr": "Lifestyle (parent topic)", "/m/032tl": "Fashion", "/m/027x7n": "Fitness",
                        "/m/02wbm": "Food", "/m/03glg": "Hobby", "/m/068hy": "Pets",
                        "/m/041xxh": "Physical attractiveness [Beauty]", "/m/07c1v": "Technology",
                        "/m/07bxq": "Tourism", "/m/07yv9": "Vehicles", "/m/098wr": "Society (parent topic)",
                        "/m/09s1f": "Business", "/m/0kt51": "Health", "/m/01h6rj": "Military", "/m/05qt0": "Politics",
                        "/m/06bvp": "Religion", "/m/01k8wb": "Knowledge", '/g/120yrv6h': 'Tourism'}
            request = api_service.channels().list(
                part="snippet,topicDetails,status",
                id=channel_url.split('/')[-1]
            )
            response = request.execute()

            # Check if the response has items
            if 'items' in response:
                channel_properties = response['items'][0]
                made_for_kids = channel_properties.get('status', {}).get('madeForKids', 'No data')
                description = channel_properties.get('snippet', {}).get('description', 'No data')
                topic = channel_properties.get('topicDetails', {}).get('topicIds', 'No data')

                if topic != 'No data':
                    topic = ", ".join([topic_id[x] for x in reversed(topic)])

                return channel_url, made_for_kids, description, topic

            else:
                return channel_url, 'No data', 'No data', 'No data', 'No data'

        def get_processed_channels(result_worksheet):
            """Collects all channels URL already processed from Result tab in set"""
            processed_channels = set()

            for row in result_worksheet.iter_rows(min_row=2, max_col=2, values_only=True):
                channel_url = row[1]
                if channel_url:
                    processed_channels.add(channel_url)

            return processed_channels

        # Initialize Language Detection
        l_detector = LanguageDetectorBuilder.from_all_languages().with_preloaded_language_models().build()

        # Initialize calculation process time
        start = time.time()
        current_iteration = 0
        max_iteration = self.channel_number

        # Create a new worksheet or load the "Results" sheet
        if "Results" in self.workbook.sheetnames:
            result_sheet = self.workbook["Results"]
        else:
            result_sheet = self.workbook.create_sheet(title="Results")
            result_sheet.append(["Placement", "Placement URL", "madeForKids", "Description", "Default Language", "Topic"])

        # Set the variable for future and get the channels already processed if any
        processed_channels = get_processed_channels(result_sheet)

        # Initialize YouTube API service
        youtube_api_service = get_youtube_api_service(self.api_entry.get())

        # Set the total steps of the charging bar
        self.charging_bar['maximum'] = self.channel_number

        for row in self.workbook['Data'].iter_rows(min_row=2, max_col=2, values_only=True):
            # Unpack the name and url from the current row
            channel_name, channel_url = row

            # increment by 1 the charging bar and update the iteration counter
            time.sleep(0.05)
            self.charging_bar["value"] = current_iteration
            current_iteration += 1

            if channel_url and channel_url not in processed_channels:

                try:
                    # get the properties of channel and append to result worksheet
                    col2, col3, col4, col6 = get_channel_properties(youtube_api_service, channel_url)

                    # Langauge detection
                    if col4 != "" and col4 != "error" and col4 != "No Data" and col4 is not None:
                        language = l_detector.detect_language_of(col4)
                        col5 = language.iso_code_639_3.name
                    else:
                        col5 = "No Data"

                    result_sheet.append([channel_name, col2, col3, col4, col5, col6])
                    # add channel url to processed channels
                    processed_channels.add(channel_url)
                    print(f"processing {channel_name} - {channel_url}")

                except Exception as e:
                    # Handle quota exceeded error
                    if "quotaExceeded" in str(e):
                        messagebox.showinfo(title="Message Box", message="Quota exceeded: Result saved in your file", icon='info')
                        break  # Exit the loop

                    # append channel data to result worksheet with error if exception raised
                    result_sheet.append([channel_name, channel_url, "error", "error", "error", "error"])
                    # add channel url to processed channels
                    processed_channels.add(channel_url)
                    print(f"Error processing {channel_name} - {channel_url}: {str(e)}")
            else:
                print(f"Already processed {channel_name} - {channel_url}")

            if self.stop_and_save_state:
                # check state of attribute, if user click the button, exit the loop
                break

            # update the time left and display it
            process_time = calculation_process_time(start, current_iteration, max_iteration)
            self.lbl_yb_channel_count.config(text=f"{process_time}")
            # update the charging bar and rest it for next iteration
            self.charging_bar.update()
            self.charging_bar["value"] = 0

        # end of for loop
        # load charging bar to max and hide it
        self.charging_bar["value"] = self.channel_number
        self.charging_bar.hide_bar()
        # save collected data in file
        self.workbook.save(self.excel_file_path.get())
        # display with message box process done
        messagebox.showinfo(title="Message Box", message="Process done: Result saved in your file", icon='info')
        # reset app for new process
        self.end_process()

    def end_process(self):
        """reset attribute and interface for potential new process cycle"""
        self.lbl_yb_channel_count.config(text="")
        self.btn_process.config(text=f"Process channels", command=self.process_channels, state=tk.DISABLED)
        self.channel_number = 0
        self.processed_channel = 0
        self.workbook = None
        self.stop_and_save_state = False
        self.lbl_file_uploaded.config(text="No file uploaded", image="", anchor=tk.CENTER)
        self.api_entry.config(state=tk.NORMAL)
        self.btn_upload.config(state=tk.NORMAL)


class Container(ttk.Frame):
    """Layout formatter

    To organize the widgets over the interface and have consistent layout it's recommended to use frame.
    Widgets are placed in frames with similar settings.
    Having standard frame with this class help to reduce line of code and improve readability

    Typical use:
        widget_container = Container(self.master, column_number, row_number, uniform_type)
    """
    def __init__(self, master, column_number, row_number, uniform_type=None):
        """
        Parameters
        ----------
        master :  Optional[Misc]
            parent of the container
        column_number : Any
            number of column of the container
        row_number : Any
            number of row of the container
        uniform_type : str
            if non-empty, empty row and column will have the same size as others
        """
        super().__init__(master)
        self.pack(fill=tk.BOTH, expand=True)
        self.columnconfigure(column_number, weight=1, uniform=uniform_type)
        self.rowconfigure(row_number, weight=1, uniform=uniform_type)


class HelpWindow(tk.Toplevel):
    """Popup window to show help instructions"""
    def __init__(self, master):
        """
        Parameters
        ----------
        master :  Optional[Misc]
            parent of the help window
        """
        super().__init__(master)
        self.title("How to get a token ?")
        self.resizable(False, False)

        self.instructions_box = TextWidget(self)


class TextWidget(tk.Text):
    """Separated class for code readability that contain the content of the HelpWindow

    Methods
    -------
    insert_text(text, tag=None)
        insert text with style in the text box
    """
    def __init__(self, master):
        """
        Parameters
        ----------
        master :  Optional[Misc]
            parent of the help text widget
        """
        super().__init__(master, wrap="none", width=95, height=20, highlightthickness=0)
        self.pack(fill=tk.BOTH, expand=True)
        self.tag_configure("bold", font=(None, 12, "bold"))
        self.tag_configure("bullet", font=(None, 10), lmargin1=20, lmargin2=40)
        self.tag_configure("large", font=(None, 14, "bold"))

        # Inserted instruction content
        self.insert_text("How to Obtain API Token for YouTube Data API v3\n\n", "large")
        self.insert_text("1. Create a Project on Google Cloud Console\n", "bold")
        self.insert_text("   - Go to the Google Cloud Console: https://console.cloud.google.com\n")
        self.insert_text("   - Click on the project dropdown at the top of the page and create a new project.\n")
        self.insert_text("   - Give your project a name and click 'Create.'\n\n")

        self.insert_text("2. Enable YouTube Data API v3\n", "bold")
        self.insert_text("   - In the Google Cloud Console, use the navigation menu at the top right of the page\n")
        self.insert_text("   - Navigate to to 'APIs & Services' and click on 'Library'.\n")
        self.insert_text("   - Search for 'YouTube Data API v3' and select it.\n")
        self.insert_text("   - Click the 'Enable' button.\n\n")

        self.insert_text("3. Create API Credentials\n", "bold")
        self.insert_text("   - In the Google Cloud Console, navigate to 'APIs & Services' > 'Credentials.'\n")
        self.insert_text("   - Click on the 'Create Credentials' dropdown and select 'API Key.'\n")
        self.insert_text("   - A dialog will appear with your API key. Copy the API key.")

        self.config(state=tk.DISABLED)

    def insert_text(self, text, tag=None):
        """Insert text with style in the text box

        Parameters
        ----------
        text :  str
            text to insert in the widget
        tag : str
            style to apply on the text
        """
        self.insert(tk.END, text, tag)


class ChargingBar(ttk.Progressbar):
    """Manages the presence of the progress bar on the interface

    Methods
    -------
    show_bar()
        display the progress bar
    hide_bar()
        hide the progress bar
    """
    def __init__(self, master):
        """
        Parameters
        ----------
        master :  Optional[Misc]
            parent of the help text widget
        """
        super().__init__(master)

    def show_bar(self):
        """Show the progress bar on the app interface"""
        self.grid(row=2, column=0, sticky="nswe", padx=20, columnspan=3)

    def hide_bar(self):
        """Hide the progress bar from the app interface"""
        self.grid_forget()


if __name__ == "__main__":
    main_app = MainApp()
    main_app.mainloop()
